from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "Leads"

headers = [
    "Lead ID", "Name", "Phone Number", "Company", "Email",
    "Lead Status",
    "Call Status", "Lead Qualification", "Conversation Summary",
    "Customer Requirements", "Objections Raised",
    "Follow-up Date", "Meeting Date & Time", "Last Contacted Timestamp"
]
sheet.append(headers)

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2F5496")
for col_idx, _ in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sample_leads = [
    ["L001", "Ravi Mehta", "+91-9876543210", "Mehta Textiles", "ravi@mehtatextiles.com",
     "Pending", "", "", "", "", "", "", "", ""],
    ["L002", "Sneha Kapoor", "+91-9823456789", "Kapoor Retail Pvt Ltd", "sneha@kapoorretail.com",
     "Pending", "", "", "", "", "", "", "", ""],
    ["L003", "Arjun Nair", "+91-9812345678", "Nair Logistics", "arjun@nairlogistics.com",
     "Opted-out", "Do Not Call", "", "Requested no further contact", "", "", "", "", ""],
    ["L004", "Priya Sharma", "+91-9845123456", "Sharma Interiors", "priya@sharmainteriors.com",
     "Completed", "Meeting Booked", "Qualified", "Interested in premium package, budget confirmed",
     "CRM for interior design leads", "Price seemed high initially, resolved with annual discount",
     "", "10-Jul-2026 3:00 PM", "05-Jul-2026 11:20 AM"],
    ["L005", "Karan Malhotra", "+91-9834567123", "Malhotra Auto Parts", "karan@malhotraauto.com",
     "Pending", "", "", "", "", "", "", "", ""],
]
for row in sample_leads:
    sheet.append(row)

widths = [8, 16, 16, 22, 26, 12, 16, 16, 30, 26, 26, 14, 20, 20]
for i, w in enumerate(widths, start=1):
    sheet.column_dimensions[get_column_letter(i)].width = w

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

sheet.freeze_panes = "A2"
wb.save("data/leads_crm.xlsx")
print("Created data/leads_crm.xlsx")
