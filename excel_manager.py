from datetime import datetime
from openpyxl import load_workbook

CRM_PATH = "data/leads_crm.xlsx"

COLUMNS = [
    "Lead ID", "Name", "Phone Number", "Company", "Email",
    "Lead Status",
    "Call Status", "Lead Qualification", "Conversation Summary",
    "Customer Requirements", "Objections Raised",
    "Follow-up Date", "Meeting Date & Time", "Last Contacted Timestamp",
]
COL_IDX = {name: i + 1 for i, name in enumerate(COLUMNS)}


def get_next_pending_lead(path: str = CRM_PATH):
    """Returns (row_number, lead_dict) for the first Pending lead, or None."""
    wb = load_workbook(path)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        status = sheet.cell(row=row, column=COL_IDX["Lead Status"]).value
        if status == "Pending":
            lead = {
                col: sheet.cell(row=row, column=idx).value
                for col, idx in COL_IDX.items()
            }
            wb.close()
            return row, lead
    wb.close()
    return None


def update_lead_record(row: int, result: dict, path: str = CRM_PATH):
    """result keys: call_status, lead_qualification, conversation_summary,
    customer_requirements, objections_raised, follow_up_date, meeting_datetime"""
    wb = load_workbook(path)
    sheet = wb.active

    field_map = {
        "Call Status": result.get("call_status", ""),
        "Lead Qualification": result.get("lead_qualification", ""),
        "Conversation Summary": result.get("conversation_summary", ""),
        "Customer Requirements": result.get("customer_requirements", ""),
        "Objections Raised": result.get("objections_raised", ""),
        "Follow-up Date": result.get("follow_up_date", ""),
        "Meeting Date & Time": result.get("meeting_datetime", ""),
        "Last Contacted Timestamp": datetime.now().strftime("%d-%b-%Y %I:%M %p"),
    }
    for col, value in field_map.items():
        sheet.cell(row=row, column=COL_IDX[col]).value = value

    call_status = result.get("call_status", "")
    if call_status in ("Meeting Booked", "Not Interested"):
        sheet.cell(row=row, column=COL_IDX["Lead Status"]).value = "Completed"
    elif call_status == "No Answer":
        sheet.cell(row=row, column=COL_IDX["Lead Status"]).value = "Pending"

    wb.save(path)
    wb.close()
